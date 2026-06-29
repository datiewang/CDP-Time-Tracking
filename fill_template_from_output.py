import os
import csv

from datetime import datetime
import re

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def find_latest_output_folder(base_dir: str) -> str:
    output_root = os.path.join(base_dir, "output")
    if not os.path.isdir(output_root):
        raise FileNotFoundError(f"'output' folder not found under {base_dir}")

    candidates = []
    for name in os.listdir(output_root):
        path = os.path.join(output_root, name)
        if not os.path.isdir(path):
            continue
        # Expect names like YYYYMMDD_HHMMSS
        try:
            datetime.strptime(name, "%Y%m%d_%H%M%S")
        except ValueError:
            continue
        candidates.append(name)

    if not candidates:
        raise FileNotFoundError("No timestamped subfolders found under 'output'.")

    latest_name = sorted(candidates)[-1]
    return os.path.join(output_root, latest_name)


def find_summary_csv(output_dir: str) -> str:
    for name in os.listdir(output_dir):
        if name.startswith("cdp_time_by_person_month") and name.endswith(".csv"):
            return os.path.join(output_dir, name)
    raise FileNotFoundError("Could not find cdp_time_by_person_month*.csv in latest output folder.")


def load_hours_by_person_month(csv_path: str) -> dict:
    mapping = {}
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            person = (row.get("Name") or "").strip()
            month = (row.get("Month") or "").strip()
            total = (row.get("TotalHours") or "").strip()
            if not person or not month or not total:
                continue
            try:
                hours = float(total)
            except ValueError:
                continue
            mapping[(person, month)] = hours
    return mapping


def _extract_month_key(cell_value) -> str | None:
    """
    Template has month labels like '4月', '5月', ... in the first column.
    Convert them to CSV-style keys like '2025/4'.
    Only accept cells that look like a standalone month label,
    not long sentences that happen to contain digits.
    """
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    if not s:
        return None
    m = re.fullmatch(r"\s*(\d+)\s*月\s*", s)
    if not m:
        return None
    month_num = int(m.group(1))
    # Standard Japanese fiscal year 2025: Months 4-12 are 2025, months 1-3 are 2026
    year = 2025 if month_num >= 4 else 2026
    return f"{year}/{month_num}"


def fill_template(template_path: str, output_dir: str, data: dict) -> str:
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is not installed. Please install it first, e.g. 'pip install openpyxl'."
        )

    wb = load_workbook(template_path)
    ws = wb.active

    # 1) Detect month rows: first column contains something like '4月', '5月', ...
    month_row_to_key: dict[int, str] = {}
    for row in range(1, ws.max_row + 1):
        month_key = _extract_month_key(ws.cell(row=row, column=1).value)
        if not month_key:
            continue
        month_row_to_key[row] = month_key

    if not month_row_to_key:
        raise RuntimeError("Could not detect any month rows in template.")

    # 2) Header row for people names is just above the first month row.
    header_row = min(month_row_to_key.keys()) - 1
    if header_row <= 0:
        raise RuntimeError("Detected month rows but could not determine header row for people.")

    # 3) Collect person names from header row, starting from column 5 onward.
    person_by_col: dict[int, str] = {}
    for col in range(5, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue
        name = str(value).strip()
        if not name:
            continue
        person_by_col[col] = name

    if not person_by_col:
        raise RuntimeError("Could not detect any person names in template header row.")

    # 4) Fill cells: for each (person, month) that exists in data, write hours.
    for col, person_name in person_by_col.items():
        # CSV uses names like 'Tetsuya KANEHIRA/ONHM'
        csv_person_name = f"{person_name}/ONHM"
        for row, month_key in month_row_to_key.items():
            key = (csv_person_name, month_key)
            hours = data.get(key)
            if hours is None:
                continue
            cell = ws.cell(row=row, column=col)
            # Only change the value; styles/formatting are preserved.
            cell.value = hours

    output_path = os.path.join(output_dir, "filled_template.xlsx")
    wb.save(output_path)
    return output_path


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))

    template_path = os.path.join(base_dir, "template.xlsx")
    if not os.path.exists(template_path):
        print(f"template.xlsx not found at: {template_path}")
        return

    try:
        latest_output_dir = find_latest_output_folder(base_dir)
    except Exception as e:
        print(str(e))
        return

    try:
        summary_csv = find_summary_csv(latest_output_dir)
    except Exception as e:
        print(str(e))
        return

    print(f"Using latest output folder: {latest_output_dir}")
    print(f"Using summary CSV: {summary_csv}")

    data = load_hours_by_person_month(summary_csv)
    if not data:
        print("No (person, month) data loaded from CSV.")
        return

    try:
        filled_path = fill_template(template_path, latest_output_dir, data)
    except Exception as e:
        print(f"Failed to fill template: {e}")
        return

    print(f"Filled workbook saved to: {filled_path}")


if __name__ == "__main__":
    main()
